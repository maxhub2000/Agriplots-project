import subprocess
import shutil
import pandas as pd
import time

import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter




# from choose_files_for_model import *
from choosing_parameters_interface import *
from editing_mod_file import *






def create_yeshuvim_with_locations(df_):
    yeshuvim_locations_df = df_[['location_id', 'YeshuvName']]
    D = {}
    for YeshuvName in yeshuvim_locations_df["YeshuvName"].unique():
        # Get all location IDs associated with the current YeshuvName
        location_ids = yeshuvim_locations_df.loc[yeshuvim_locations_df['YeshuvName'] == YeshuvName]["location_id"].tolist()
        if YeshuvName not in D:
            D[YeshuvName] = set(location_ids)  # Create a set with all locations for the YeshuvName
        else:
            D[YeshuvName].update(location_ids)  # Update the set with additional location_ids if any   
    return D
  
def adjust_energy_consumption_by_yeshuv(energy_consumption_by_yeshuv_, relevant_yeshuvim_):
    energy_consumption_by_yeshuv_ = energy_consumption_by_yeshuv_.drop(['yeshuv_symbol'], axis=1) # remove yeshuv_symbol column
    # creating energy consumption by yeshuv for only yeshuvim in inputed df (relevant_yeshuvim_),
    # if yeshuv not in original energy_consumption_by_yeshuv, their yearly consumption will be 0
    adjusted_df = pd.DataFrame(columns=['yeshuv_name', 'yearly energy consumption'])
    for yeshuv in relevant_yeshuvim_:
        if yeshuv in energy_consumption_by_yeshuv_["yeshuv_name"].tolist():
            # if yeshuv in input df, get it's yeshuv name and yearly energy consumption in the "row" var, then append the row to the adjusted_df
            row = energy_consumption_by_yeshuv_.loc[energy_consumption_by_yeshuv_['yeshuv_name'] == yeshuv]
            adjusted_df = adjusted_df.append(row, ignore_index = True)
        else:
            # otherwise, add the yeshuv name with yearly energy consumption of 0, so that the model wouldn't choose locations from that yeshuv
            row = {'yeshuv_name': yeshuv, 'yearly energy consumption': 0.00}
            adjusted_df = adjusted_df.append(row, ignore_index = True)
    return adjusted_df

def create_machozot_with_locations(df_):
    machozot_locations_df = df_[['location_id', 'Machoz']]
    D = {}
    for Machoz in machozot_locations_df["Machoz"].unique():
        # Get all location IDs associated with the current Machoz
        location_ids = machozot_locations_df.loc[machozot_locations_df['Machoz'] == Machoz]["location_id"].tolist()
        if Machoz not in D:
            D[Machoz] = set(location_ids)  # Create a set with all locations for the Machoz
        else:
            D[Machoz].update(location_ids)  # Update the set with additional location_ids if any   
    return D

def adjust_energy_consumption_by_machoz(energy_consumption_by_machoz_, relevant_machozot_):
    # creating energy consumption by machoz for only machozot in inputed df (relevant_machozot_),
    # if machoz not in original energy_consumption_by_machoz, their yearly consumption will be 0
    adjusted_df = pd.DataFrame(columns=['machoz', 'yearly energy consumption'])
    for machoz in relevant_machozot_:
        if machoz in energy_consumption_by_machoz_["machoz"].tolist():
            # if machoz in input df, get it's machoz name and yearly energy consumption in the "row" var, then append the row to the adjusted_df
            row = energy_consumption_by_machoz_.loc[energy_consumption_by_machoz_['machoz'] == machoz]
            adjusted_df = adjusted_df.append(row, ignore_index = True)
        else:
            # otherwise, add the machoz name with yearly energy consumption of 0, so that the model wouldn't choose locations from that machoz
            row = {'machoz': machoz, 'yearly energy consumption': 0.00}
            adjusted_df = adjusted_df.append(row, ignore_index = True)
    return adjusted_df

def sort_df_by_list_order(df_, list_order, column_name):
    """
    Sort a DataFrame by the order of values in a list.

    Parameters:
    df (pd.DataFrame): The DataFrame to sort.
    list_order (list): A list of values specifying the desired order.
    column_name (str): The column in df to be sorted based on the list_order.

    Returns:
    pd.DataFrame: A new DataFrame sorted according to the list_order.
    """
    # Convert the specified column to a categorical type with the provided order
    df_[column_name] = pd.Categorical(df_[column_name], categories=list_order, ordered=True)
    # Sort the DataFrame by the specified column
    df_sorted = df_.sort_values(column_name).reset_index(drop=True)
    return df_sorted

def create_eshkolot_with_locations(df_):
    eshkolot_locations_df = df_[['location_id', 'eshkol']]
    D = {}
    eshkolot_lst = eshkolot_locations_df["eshkol"].unique()
    eshkolot_lst = sorted(eshkolot_lst)
    #print("eshkolot_lst",eshkolot_lst)
    for eshkol in eshkolot_lst:
        # Get all location IDs associated with the current eshkol
        location_ids = eshkolot_locations_df.loc[eshkolot_locations_df['eshkol'] == eshkol]["location_id"].tolist()
        if eshkol not in D:
            D[eshkol] = set(location_ids)  # Create a set with all locations for the eshkol
        else:
            D[eshkol].update(location_ids)  # Update the set with additional location_ids if any   
    return D

def adjust_energy_division_between_eshkolot(energy_division_between_eshkolot_, relevant_eshkolot_):
    # creating energy division between eshkolot for only eshkolot in inputed df (relevant_eshkolot_),
    # if eshkol not in original energy_division_between_eshkolot_, their percentage_of_energy_output will be 0
    adjusted_df = pd.DataFrame(columns=['eshkol', 'percentage_of_energy_output'])
    for eshkol in relevant_eshkolot_:
        if eshkol in energy_division_between_eshkolot_["eshkol"].tolist():
            # if eshkol in input df, get it's eshkol number and percentage of energy output in the "row" var, then append the row to the adjusted_df
            row = energy_division_between_eshkolot_.loc[energy_division_between_eshkolot_['eshkol'] == eshkol]
            adjusted_df = adjusted_df.append(row, ignore_index = True)
        else:
            # otherwise, add the eshkol num with percentage of energy output of 0, so that the model wouldn't choose locations from that eshkol
            row = {'eshkol': eshkol, 'percentage_of_energy_output': 0.00}
            adjusted_df = adjusted_df.append(row, ignore_index = True)
    return adjusted_df

def prepare_data(df_, energy_consumption_by_yeshuv, energy_division_between_eshkolot, energy_consumption_by_machoz, total_potential_revenue_before_PV_of_full_dataset):
    fix_energy_production = df_['Energy production (fix) mln kWh/year'].tolist()
    area_in_dunam = df_['Dunam'].tolist()
    influence_on_crops = df_['Average influence of PV on crops'].tolist()
    potential_revenue_before_PV = df_['Potential revenue from crops before PV, mln NIS'].tolist()
    num_locations = len(fix_energy_production)


    yeshuvim_with_locations = create_yeshuvim_with_locations(df_)
    # takes only yeshuvim that appear in the current dataset; values are unique since yeshuvim_with_locations is a dictionary
    relevant_yeshuvim = list(yeshuvim_with_locations.keys())
    num_yeshuvim = len(relevant_yeshuvim)
    # adjust energy consumptions so that it will only include yeshuvim from the dataset, and then convert it to a list
    energy_consumption_by_yeshuv = adjust_energy_consumption_by_yeshuv(energy_consumption_by_yeshuv, relevant_yeshuvim)
    energy_consumption_by_yeshuv = energy_consumption_by_yeshuv['yearly energy consumption'].tolist()


    machozot_with_locations = create_machozot_with_locations(df_)
    # takes only machozot that appear in the current dataset; values are unique since machozot_with_locations is a dictionary
    relevant_machozot = list(machozot_with_locations.keys())
    num_machozot = len(relevant_machozot)
    # adjust energy consumptions so that it will only include machozot from the dataset, and then convert it to a list
    energy_consumption_by_machoz = adjust_energy_consumption_by_machoz(energy_consumption_by_machoz, relevant_machozot)
    energy_consumption_by_machoz = energy_consumption_by_machoz['yearly energy consumption'].tolist()


    #eshkolot_with_locations = df_[['location_id', 'eshkol']][df_['eshkol'] == -1]
    eshkolot_with_locations = create_eshkolot_with_locations(df_)
    #print("eshkolot_with_locations:\n", eshkolot_with_locations)
    # takes only eshkolot that appear in the current dataset; values are unique since eshkolot_with_locations is a dictionary
    relevant_eshkolot = list(eshkolot_with_locations.keys())
    #print("relevant_eshkolot:\n",relevant_eshkolot)
    num_eshkolot = len(relevant_eshkolot)
    # adjust energy_division_between_eshkolot so that it will only include eshkolot from the dataset, and then convert it to a list
    energy_division_between_eshkolot = adjust_energy_division_between_eshkolot(energy_division_between_eshkolot, relevant_eshkolot)
    energy_division_between_eshkolot = energy_division_between_eshkolot['percentage_of_energy_output'].tolist()


    return {
        "num_locations" : num_locations, #need to be first here, otherwise there will be bug in the mod file
        "fix_energy_production" : fix_energy_production,
        "influence_on_crops" : influence_on_crops,
        "potential_revenue_before_PV" : potential_revenue_before_PV,
        "total_potential_revenue_before_PV_of_full_dataset" : total_potential_revenue_before_PV_of_full_dataset,
        "area_in_dunam" : area_in_dunam,
        "yeshuvim_with_locations" : yeshuvim_with_locations,
        "num_yeshuvim" : num_yeshuvim,
        "energy_consumption_by_yeshuv" : energy_consumption_by_yeshuv,
        "machozot_with_locations" : machozot_with_locations,
        "num_machozot" : num_machozot,
        "energy_consumption_by_machoz" : energy_consumption_by_machoz,
        "eshkolot_with_locations" : eshkolot_with_locations,
        "num_eshkolot" : num_eshkolot,
        "energy_division_between_eshkolot" : energy_division_between_eshkolot
    }

def write_dat_file(dat_file, data, params):
    yeshuvim_with_locations = data.pop("yeshuvim_with_locations")
    machozot_with_locations = data.pop("machozot_with_locations")
    eshkolot_with_locations = data.pop("eshkolot_with_locations")
    with open(dat_file, 'w') as f:
        for dict in [params, data]:
            for key, val in dict.items():
                f.write(f"{key} = {val};\n")

        f.write("S = [\n")
        for yeshuv, locations in yeshuvim_with_locations.items():
            f.write(f"{locations},\n")
        f.write("];\n")
        
        f.write("M = [\n")
        for machoz, locations in machozot_with_locations.items():
            f.write(f"{locations},\n")
        f.write("];")

        f.write("E = [\n")
        for eshkol, locations in eshkolot_with_locations.items():
            f.write(f"{locations},\n")
        f.write("];")

def solve_opl_model(mod_file, dat_file, output_file=None):
    oplrun_path = "oplrun"
    if not shutil.which(oplrun_path):
        print(f"{oplrun_path} not found in PATH. Make sure CPLEX Optimization Studio is installed and oplrun is in the PATH.")
        return

    command = [oplrun_path, mod_file, dat_file]
    try:
        result = subprocess.run(command, capture_output=True, text=True)
        if output_file:
            with open(output_file, 'w') as f:
                if result.returncode == 0:
                    f.write("Solution found:\n")
                    f.write(result.stdout)
                else:
                    f.write("Error in solving the model:\n")
                    f.write("Return code: " + str(result.returncode) + "\n")
                    f.write("Standard Output: " + result.stdout + "\n")
                    f.write("Standard Error: " + result.stderr + "\n")
            print(f"Output saved to {output_file}")
        else:
            if result.returncode == 0:
                print("Solution found:")
                print(result.stdout)
            else:
                print("Error in solving the model:")
                print("Return code:", result.returncode)
                print("Standard Output:", result.stdout)
                print("Standard Error:", result.stderr)
        return result.stdout
    except Exception as e:
        print(f"Error running oplrun: {e}")

def modify_influence_on_crops(df_, synthetic_values_of_influence_on_crops_path):
    # prepare dictionary that maps for each crop how much it's influenced by installing PV
    influence_on_crops_data = pd.read_excel(synthetic_values_of_influence_on_crops_path)
    influence_on_crops_dict = influence_on_crops_data.set_index("AnafSub")["Average influence"].to_dict()
    # maps average influence on crops to each AnafSub according to the influence_on_crops_dict (like Vlookp)
    modified_influence_on_crops = df_['AnafSub'].map(influence_on_crops_dict).tolist()
    df_["Average influence of PV on crops"] = modified_influence_on_crops
    return df_

def remove_rows_with_missing_values(df_):
    # remove rows with nan values (Ideally should find a better way to handle those nan values later on)
    df_ = df_.dropna(subset=['Energy production (fix) mln kWh/year',
                           'Average influence of PV on crops',
                           'Total revenue, mln NIS',
                           'Dunam'])
    return df_


def remove_rows_with_non_feasible_locations(df_):
    # remove rows with non feasible locations for installing PV's
    df_ = df_[df_['Feasability to install PVs?'] != 0]
    return df_


def add_eshkolot_to_dataset(df_, yeshuvim_in_eshkolot_):
    # Perform a left join to add the 'eshkol' column from yeshuvim_in_eshkolot_ to df_
    df_ = pd.merge(df_, yeshuvim_in_eshkolot_, on='YeshuvName', how='left')
    # Fill missing values (if a YeshuvName is not found in yeshuvim_in_eshkolot_, set 'eshkol' to -1)
    df_['eshkol'].fillna(-1, inplace=True)
    # Optionally, convert 'eshkol' to integer if needed
    df_['eshkol'] = df_['eshkol'].astype(int)
    #remove rows that has eshkol -1, meaning rows that their yeshuv doesn't have an eshkol
    df_ = df_[df_['eshkol'] != -1]
    return df_

def raw_output_to_df(opl_raw_output_):
    # parsing the result output to extract the required values
    main_results = pd.DataFrame()
    locations_with_installed_PVs = ""  # Variable to store multiple lines of output
    energy_produced_per_eshkol = ""  # Variable to store multiple lines of output
    capture_main_results = False
    capture_locations_with_installed_PVs_results = False  # Flag to start capturing excel output
    capture_energy_produced_per_eshkol_results = False  # Flag to start capturing excel output

    for line in opl_raw_output_.splitlines():
        # Detect the start of the excel output block
        if "Results for excel output file:" in line:
            # print("in first condition")
            capture_main_results = True  # Start capturing from the next line
            continue  # Skip the current line


        if "Locations with installed PV's:" in line:
            # print("in second condition")
            capture_main_results = False
            capture_locations_with_installed_PVs_results = True
            continue


        if capture_main_results:
            # print("in third condition")
            if line.strip() == "Locations with installed PV's:":  # Stop capturing when we hit an empty line (or define another end condition)
                capture_locations_with_installed_PVs_results = False
                capture_energy_produced_per_eshkol_results = True
                continue
            splitted_line = line.strip().split(": ")
            metric, value = splitted_line 
            main_results[metric] = [value]
        
    
        
        if capture_locations_with_installed_PVs_results:
            # print("in third condition")
            if line.strip() == "Energy produced per Eshkol:":  # Stop capturing when we hit an empty line (or define another end condition)
                capture_locations_with_installed_PVs_results = False
                capture_energy_produced_per_eshkol_results = True
                continue
            locations_with_installed_PVs += line + "\n"  # Append the line to the result string

        # Capture subsequent lines after the keyword is detected
        if capture_energy_produced_per_eshkol_results:
            # print("in fourth condition")
            if line.strip() == "End of Results for excel output file":  # Stop capturing when we hit an empty line (or define another end condition)
                break
            energy_produced_per_eshkol += line + "\n"  # Append the line to the result string

    locations_with_installed_PVs = locations_with_installed_PVs.strip()
    locations_with_installed_PVs = locations_with_installed_PVs.split("\n")  # Splitting at the \n delimiter

    energy_produced_per_eshkol = energy_produced_per_eshkol.strip()
    energy_produced_per_eshkol = energy_produced_per_eshkol.split("\n")  # Splitting at the \n delimiter

    locations_with_installed_PVs_results = model_results_to_df(locations_with_installed_PVs)
    energy_produced_per_eshkol_results = model_results_to_df(energy_produced_per_eshkol)
    print("locations_with_installed_PVs_results:\n",locations_with_installed_PVs_results)
    print("energy_produced_per_eshkol_results:\n",energy_produced_per_eshkol_results)

    print("main_results:\n", main_results)
    df_results = [main_results, locations_with_installed_PVs_results, energy_produced_per_eshkol_results]
    # Return the captured results, including the excel output block
    return df_results

def model_results_to_df(excel_output_results):
    # Split the first element to get the column names
    column_names = excel_output_results[0].split(',')
    # Split each remaining element in the list to get the data rows
    rows = [row.split(',') for row in excel_output_results[1:]]
    # Create a pandas DataFrame using the rows and columns
    df_results = pd.DataFrame(rows, columns = column_names)
    return df_results

def output_opl_results_to_excel(df_input, df_opl_results, model_params, output_path):
    main_results_df, locations_with_installed_PVs_results, energy_produced_per_eshkol_results = df_opl_results

    relevant_columns_from_input = ["location_id", "OBJECTID", "AnafSub", "YeshuvName", "Machoz", "eshkol", "Potential revenue from crops before PV, mln NIS", "Potential revenue from crops after PV, mln NIS"]
    relevant_df_from_input = df_input[relevant_columns_from_input]
    # convert "OBJECTID" and "Location" column to int, so the merge will be successful
    relevant_df_from_input['OBJECTID'] = relevant_df_from_input['OBJECTID'].astype('int')
    locations_with_installed_PVs_results['location_id'] = locations_with_installed_PVs_results['location_id'].astype('int')
    # convert "area in dunam used" column to numeric
    locations_with_installed_PVs_results["area in dunam used"] = pd.to_numeric(locations_with_installed_PVs_results["area in dunam used"], errors="coerce")
    # left join installed locations from result of opl model to columns from input dataset
    merged_data = pd.merge(locations_with_installed_PVs_results, relevant_df_from_input, on="location_id", how="left")    
    # exporting results to excel
    area_used_per_machoz = merged_data[["Machoz", "area in dunam used"]].groupby("Machoz", as_index=False)["area in dunam used"].sum()
    #print("area_used_per_machoz:\n",area_used_per_machoz)
    area_used_per_anafSub = merged_data[["AnafSub", "area in dunam used"]].groupby("AnafSub", as_index=False)["area in dunam used"].sum()
    #print("area_used_per_AnafSub:\n",area_used_per_anafSub)
    if model_params["total_area_upper_bound"] == 1e12:
        model_params["total_area_upper_bound"] = "No Upper Bound"
    model_params_df = pd.DataFrame([model_params])
    #print(model_params_df)

    print(f"Excel output saved to {output_path}")
    merged_data.to_excel(output_path)

    full_results = [main_results_df, model_params_df, energy_produced_per_eshkol_results, area_used_per_machoz, area_used_per_anafSub]
    export_full_output_to_excel(full_results)


# Helper function to apply styling in output excel results
def style_range(ws, start_cell, end_cell, alignment=None, fill=None, font=None):
    for row in ws[start_cell:end_cell]:
        for cell in row:
            if alignment:
                cell.alignment = alignment
            if fill:
                cell.fill = fill
            if font:
                cell.font = font


def export_full_output_to_excel(full_results):

    main_results, model_params, energy_produced_per_eshkol_results, area_used_per_machoz, area_used_per_anafSub = full_results

    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # Apply column widths
    column_widths = [30, 30, 30, 30, 30, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Write the "Model Results" table with formatting
    ws.append(["Model Results"])
    style_range(
        ws,
        "A1", "F1",
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        fill=PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        font=Font(bold=True, size=16)
    )
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 30  # Increase row height for better visibility

    ws.append([
        "Total energy produced in mln", "Total area (in dunam) used",
        "Gini Coefficient value", "Potential revenue before installing PV's",
        "Potential revenue after installing PV's", "Percentage change in revenue"
    ])
    style_range(
        ws,
        "A2", "F2",
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        fill=PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        font=Font(bold=True, size=14)
    )
    ws.row_dimensions[2].height = 40  # Adjust height for header row

    for row in main_results.itertuples(index=False):
        ws.append(row)

    # Add spacing
    ws.append([])

    # Write the "Model Parameters (input)" tables
    ws.append(["Model Parameters (input)"])
    style_range(
        ws,
        f"A{ws.max_row}", f"F{ws.max_row}",
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        fill=PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        font=Font(bold=True, size=16)
    )
    ws.merge_cells(f"A{ws.max_row}:C{ws.max_row}")
    ws.row_dimensions[ws.max_row].height = 30  # Increase row height

    ws.append([
        "Percentage change in revenue lower bound", "Total area upper bound",
        "Gini Coefficient upper bound"
    ])
    style_range(
        ws,
        f"A{ws.max_row}", f"F{ws.max_row}",
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        fill=PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        font=Font(bold=True, size=14)
    )
    ws.row_dimensions[ws.max_row].height = 40  # Adjust height for parameter row

    for row in model_params.itertuples(index=False):
        ws.append(row)

    # Add similar logic for other parameter DataFrames
    ws.append([])
    ws.append(["Energy produced per Eshkol"])
    style_range(
        ws,
        f"A{ws.max_row}", f"B{ws.max_row}",
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        fill=PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        font=Font(bold=True, size=16)
    )
    ws.merge_cells(f"A{ws.max_row}:B{ws.max_row}")
    ws.row_dimensions[ws.max_row].height = 25

    ws.append(["Eshkol num", "Energy Produced"])
    style_range(
        ws,
        f"A{ws.max_row}", f"B{ws.max_row}",
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        fill=PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        font=Font(bold=True, size=14)
    )

    for row in energy_produced_per_eshkol_results.itertuples(index=False):
        ws.append(row)

    ws.append([])
    ws.append(["Area user per Machoz"])
    style_range(
        ws,
        f"A{ws.max_row}", f"B{ws.max_row}",
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        fill=PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        font=Font(bold=True, size=16)
    )
    ws.merge_cells(f"A{ws.max_row}:B{ws.max_row}")
    ws.row_dimensions[ws.max_row].height = 25

    ws.append(["Machoz", "Area in dunam Used"])
    style_range(
        ws,
        f"A{ws.max_row}", f"B{ws.max_row}",
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        fill=PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        font=Font(bold=True, size=14)
    )

    for row in area_used_per_machoz.itertuples(index=False):
        ws.append(row)

    ws.append([])
    ws.append(["Area user per AnafSub"])
    style_range(
        ws,
        f"A{ws.max_row}", f"B{ws.max_row}",
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        fill=PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        font=Font(bold=True, size=16)
    )
    ws.merge_cells(f"A{ws.max_row}:B{ws.max_row}")
    ws.row_dimensions[ws.max_row].height = 25

    ws.append(["AnafSub", "Area in dunam Used"])
    style_range(
        ws,
        f"A{ws.max_row}", f"B{ws.max_row}",
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        fill=PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        font=Font(bold=True, size=14)
    )

    for row in area_used_per_anafSub.itertuples(index=False):
        ws.append(row)

    # Adjust alignment for all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Save the file
    full_results_output_path = "full_results.xlsx"
    wb.save(full_results_output_path)
    print(f"Excel output saved to {full_results_output_path}")

    os.startfile(full_results_output_path)




def main():
    # Display the file selection interface
    # opl_model_file, dat_file = select_files()
    
    # if not opl_model_file or not dat_file:
    #     print("Please select both the .mod and .dat files to proceed.")
    #     return
    
    output_file = 'output.txt'

    # Save start time of running of the code
    start_time_code = time.time()
    # File paths and parameters
    opl_model_file, dat_file, output_file = 'Agriplots.mod', 'Agriplots.dat', 'output.txt'
    # opl_model_file = 'models/basic_model_with_gini/Agriplots_basic_with_gini.mod'
    #opl_model_file = 'advanced_model_with_gini_in_objective_function/Agriplots_advanced_model_with_gini_in_objective_function.mod'
    
    # Call the function to get the selected .mod file path
    # opl_model_file, dataset_path = open_file_selector()
    dataset_path = 'Agriplots_final - Full data.xlsx'
    dataset_path = 'Agriplots dataset - 1,000 rows.xlsx'

    
    df_dataset = pd.read_excel(dataset_path) # Read dataset from Excel
    print("number of rows in full dataset :", len(df_dataset))
    total_potential_revenue_before_PV_of_full_dataset = df_dataset["Potential revenue from crops before PV, mln NIS"].sum() #parameter to pass later on
    print("total_potential_revenue_before_PV:",total_potential_revenue_before_PV_of_full_dataset)
    df_dataset = remove_rows_with_missing_values(df_dataset)
    print("number of rows in dataset after removing rows with missing values:", len(df_dataset))
    df_dataset = remove_rows_with_non_feasible_locations(df_dataset)
    print("number of rows in dataset after removing rows with non feasible locations:", len(df_dataset))
    # modify influence on crops column based on synthetic values
    df_dataset = modify_influence_on_crops(df_dataset, 'Average influence of PV on crops - synthetic values.xlsx')
    # import energy consumptions by yeshuv and by machoz 
    energy_consumption_by_yeshuv = pd.read_excel("energy_consumption_by_yeshuv-average_consumption_times_population_per_yeshuv.xlsx")
    energy_consumption_by_machoz = pd.read_excel("energy_consumption_by_machoz_aggregated_from_yeshuvim.xlsx", sheet_name = "energy consumption by machoz")
    # import datasets relevant for using eshkolot in the model
    #yeshuvim_in_eshkolot = pd.read_excel('yeshuvim_in_eshkolot.xlsx')
    yeshuvim_in_eshkolot = pd.read_excel('yeshuvim_in_eshkolot_modified_to_match_dataset.xlsx')
    yeshuvim_in_eshkolot.rename(columns = {'eshkol_2021':'eshkol'}, inplace = True) # rename column in the df
    energy_division_between_eshkolot = pd.read_excel('energy_division_between_eshkolot-synthetic_values.xlsx')
    # add eshkolot to dataset based on yeshuvim_in_eshkolot, and also remove rows that their yeshuv doesn't have an eshkol
    df_dataset = add_eshkolot_to_dataset(df_dataset, yeshuvim_in_eshkolot)
    print("number of rows in dataset after adding eshkolot:", len(df_dataset))

    #print("df_dataset after adding eshkolot:\n", df_dataset)
    #df_dataset.to_excel("df_dataset_after_adding_eshkolot.xlsx")
    # create location_id column in df_dataset that's based on index of the df
    df_dataset = df_dataset.reset_index() # reset index of the df before creating the new column, since rows were removed earlier
    df_dataset["location_id"] = df_dataset.index + 1 
    #print("df_dataset[location_id]\n", df_dataset["location_id"])
    # parameters of the model
    params = {
        "allowed_loss_from_influence_on_crops_percentage": 0.9,
        "total_area_upper_bound": 30000.00,
        "G_max" : 0.05
        # "G_max" : 1.00
    }


    user_input_params = activate_interface()
    trillion = 1e12 # float form of trillion that's also compatible with .mod and .dat files
    if user_input_params[0]:
        params["total_area_upper_bound"] = user_input_params[0]
    else:
        params["total_area_upper_bound"] = trillion # simulating infinity to show it won't be used in the model
    params["allowed_loss_from_influence_on_crops_percentage"] = user_input_params[1]/100
    removed_constraints = user_input_params[2]

    template_file_path = "edit_mod_file/template.mod"
    edited_file_path = "edit_mod_file/edited.mod"

    constraints_to_comment_texts = {
        "total_area_constraint" : "Constraint for an upper bound of the total area used by installed PV's",
        "revenue_change_constraint" : "Constraint for the revenue change in percentage as a result of installing the PV's and influencing the crops, lower bounded by an inputed threshold",
        "energy_production_per_yeshuv_constraint" : "Constraint for the total energy production of each yeshuv, upper bounded by the energy consumption of each yeshuv",
        "energy_production_per_machoz_constraint" : "Constraint for the total energy production of each machoz, upper bounded by the energy consumption of each machoz",
        "name_space_gini_constraint1": "Linearized Gini coefficient constraint (only for i < j)",
        "name_space_gini_constraint2": "Gini constraint (now summing only over i < j)",
    }

    def remove_constraints_from_model(removed_constraints, template_file_path, edited_file_path, constraints_to_comment_texts):
        comments_in_mod_file = []
        for removed_constraint in removed_constraints:
            comments_in_mod_file.append(constraints_to_comment_texts[removed_constraint])
        comment_multiple_sections_in_mod(template_file_path, edited_file_path, comments_in_mod_file)
        print("\n")
        print("The following constraints were removed from the model:")
        for removed_constraint in removed_constraints:
            print(removed_constraint)
        print("\n")


    remove_constraints_from_model(removed_constraints, template_file_path, edited_file_path, constraints_to_comment_texts)

    time.sleep(10)

    opl_model_file = edited_file_path

    
    #params["G_max"] = 0.05


    # get needed relevant data for running the model, in addition to the parameters (params)
    data = prepare_data(df_dataset, energy_consumption_by_yeshuv, energy_division_between_eshkolot, energy_consumption_by_machoz, total_potential_revenue_before_PV_of_full_dataset)
    # write data and params to .dat file
    write_dat_file(dat_file, data, params)

    # Save end time of running of the code except for the opl model
    end_time_code = time.time()
    run_time_code_before_opl_model = end_time_code - start_time_code
    print("\nrun_time_code_before_opl_model:",run_time_code_before_opl_model, "\n")
    # Save start time of running of opl model
    start_time_opl_model = time.time()
    # Solve the OPL model and put the opl output in the opl_raw_output variable
    opl_raw_output = solve_opl_model(opl_model_file, dat_file, output_file)
    # Save end time of running of opl model
    end_time_opl_model = time.time()
    run_time_opl_model = end_time_opl_model - start_time_opl_model
    print("\nrun_time_opl_model:",run_time_opl_model,"\n")

    # converting the raw output of the model to a dataframe with the needed results
    df_results = raw_output_to_df(opl_raw_output)
    #print("model_results:\n", df_results)
    # output the final results to excel file
    final_results_excel_output_path = 'final_opl_results.xlsx'
    output_opl_results_to_excel(df_dataset, df_results, params, final_results_excel_output_path)

    time.sleep(1)

    # Specify the path to your Excel file
    file_path = "final_opl_results.xlsx"
    #os.startfile(file_path)



if __name__ == "__main__":
    main()
