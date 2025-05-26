import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
from utils import measure_time


@measure_time
def output_opl_results_to_excel(df_dataset_, df_opl_results, model_params, installation_decisions_output_path, final_results_output_path, decision_variables_type, objective_function_type, main_constraint):
    main_results_df, installed_PVs_results, energy_produced_per_eshkol = df_opl_results
    installation_decisions = output_installation_decisions_results_to_excel(df_dataset_, installed_PVs_results, installation_decisions_output_path)
    energy_produced_per_machoz = installation_decisions[["Machoz", "Energy units Produced in mln"]].groupby("Machoz", as_index=False)["Energy units Produced in mln"].sum()
    area_used_per_machoz = installation_decisions[["Machoz", "area in dunam used"]].groupby("Machoz", as_index=False)["area in dunam used"].sum()
    area_used_per_anafSub = installation_decisions[["AnafSub", "area in dunam used"]].groupby("AnafSub", as_index=False)["area in dunam used"].sum()
    energy_produced_per_eshkol["Energy Produced"] = pd.to_numeric(energy_produced_per_eshkol["Energy Produced"], errors="coerce")
    main_results_df = main_results_df.apply(pd.to_numeric, errors='coerce')
    if model_params["total_area_upper_bound"] == 1e12:
        model_params["total_area_upper_bound"] = "No Upper Bound"
    model_params_df = pd.DataFrame([model_params])
    full_results = [main_results_df, model_params_df, energy_produced_per_eshkol, energy_produced_per_machoz, area_used_per_machoz, area_used_per_anafSub]
    output_final_results_to_excel(full_results, final_results_output_path, decision_variables_type, objective_function_type, main_constraint)

def output_installation_decisions_results_to_excel(df_dataset_, installed_PVs_results, installation_decisions_output_path):
    relevant_columns_from_input = ["location_id", "OBJECTID", "AnafSub", "YeshuvName", "Machoz", "eshkol", "Potential revenue from crops before PV, mln NIS", "Potential revenue from crops after PV, mln NIS"]
    if "OBJECTID" not in list(df_dataset_.columns):
        relevant_columns_from_input.remove("OBJECTID")
    relevant_df_from_input = df_dataset_[relevant_columns_from_input]
    # convert "OBJECTID" and "Location" column to int, so the merge will be successful
    if "OBJECTID" in list(df_dataset_.columns):
        relevant_df_from_input['OBJECTID'] = relevant_df_from_input['OBJECTID'].astype('int')
    installed_PVs_results['location_id'] = installed_PVs_results['location_id'].astype('int')
    # convert "Energy units Produced in mln" and "area in dunam used" columns to numeric
    installed_PVs_results["Energy units Produced in mln"] = pd.to_numeric(installed_PVs_results["Energy units Produced in mln"], errors="coerce")
    installed_PVs_results["area in dunam used"] = pd.to_numeric(installed_PVs_results["area in dunam used"], errors="coerce")
    # left join installed locations from result of opl model to columns from input dataset
    merged_data = pd.merge(installed_PVs_results, relevant_df_from_input, on="location_id", how="left")    
    # exporting results to excel
    print(f"installation decisions results saved to {installation_decisions_output_path}")
    merged_data.to_excel(installation_decisions_output_path)
    return merged_data

def output_final_results_to_excel(full_results, final_results_output_path, decision_variables_type, objective_function_type, main_constraint):
    main_results, model_params, energy_produced_per_eshkol, energy_produced_per_machoz, area_used_per_machoz, area_used_per_anafSub = full_results
    # Create excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    set_column_widths(ws, [30]*6)
    # construct model results title
    model_results_title = construct_model_results_title(decision_variables_type, objective_function_type, main_constraint)
    # Main results Column titles assigned to variables
    results_column_titles = [
        "Total energy produced in mln", "Total area (in dunam) used",
        "Remaining percentage of revenue", "Total installation cost",
        "Potential revenue before installing PV's", "Potential revenue after installing PV's"
    ]
    if "Gini Coefficient value" in main_results.columns:
        results_column_titles.append("Gini Coefficient value")
    # Parameters column titles assigned to variables
    parameters_column_titles = [
        "Total energy produced lower bound", "Total area upper bound",
        "Remaining percentage of revenue lower bound", "Total installation cost upper bound"
    ]
    if "G_max" in model_params.columns:
        parameters_column_titles.append("Gini Coefficient upper bound")
    # Additional results Column titles assigned to variables
    energy_per_eshkol_column_titles = ["Eshkol num", "Energy Produced"]
    energy_per_machoz_column_titles = ["Machoz", "Energy Produced"]
    area_per_machoz_column_titles = ["Machoz", "Area in dunam Used"]
    area_per_anafsub_column_titles = ["AnafSub", "Area in dunam Used"]
    # Section calls
    write_section(ws, model_results_title, results_column_titles, main_results)
    write_section(ws, "Model Parameters (input)", parameters_column_titles, model_params)
    write_section(ws, "Energy produced per Eshkol", energy_per_eshkol_column_titles, energy_produced_per_eshkol)
    write_section(ws, "Energy produced per Machoz", energy_per_machoz_column_titles, energy_produced_per_machoz)
    write_section(ws, "Area used per Machoz", area_per_machoz_column_titles, area_used_per_machoz)
    write_section(ws, "Area used per AnafSub", area_per_anafsub_column_titles, area_used_per_anafSub)
    # Align all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # Save workbook and open newly created excel file
    wb.save(final_results_output_path)
    print(f"final results saved to {final_results_output_path}")
    os.startfile(final_results_output_path)

def style_range(ws, start_cell, end_cell, alignment=None, fill=None, font=None):
    for row in ws[start_cell:end_cell]:
        for cell in row:
            if alignment:
                cell.alignment = alignment
            if fill:
                cell.fill = fill
            if font:
                cell.font = font

def set_column_widths(ws, widths):
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

def write_section_title(ws, title, start_col="A", end_col="F", font_size=16, fill_color="FFFF00"):
    ws.append([title])
    row = ws.max_row
    style_range(
        ws, f"{start_col}{row}", f"{end_col}{row}",
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        fill=PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid"),
        font=Font(bold=True, size=font_size)
    )
    ws.merge_cells(f"{start_col}{row}:{end_col}{row}")
    ws.row_dimensions[row].height = 30

def write_dataframe_table(ws, column_titles, dataframe, merge_cols=2, header_font_size=14):
    ws.append(column_titles)
    row = ws.max_row
    style_range(
        ws, f"A{row}", f"{chr(64 + merge_cols)}{row}",
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        fill=PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        font=Font(bold=True, size=header_font_size)
    )
    ws.row_dimensions[row].height = 40
    for row_data in dataframe.itertuples(index=False):
        ws.append(row_data)

def write_section(ws, title, column_titles, dataframe):
    """Writes a full section: title + header row + data rows (center-aligned)."""
    # Write section title using existing helper
    write_section_title(ws, title, end_col=chr(64 + len(column_titles)))

    # Write header row
    ws.append(column_titles)
    row = ws.max_row
    style_range(
        ws,
        f"A{row}", f"{chr(64 + len(column_titles))}{row}",
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        fill=PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        font=Font(bold=True, size=14)
    )
    ws.row_dimensions[row].height = 40

    # Append data rows
    for row_data in dataframe.itertuples(index=False):
        ws.append(row_data)

    ws.append([])  # spacing row

def construct_model_results_title(decision_variables_type, objective_function_type, main_constraint):
    sheet_title = ""
    model_type_text = "Binary" if decision_variables_type == "binary decision variables" else "Continuous"
    objective_function_text = objective_function_type.capitalize()
    constraint_text = main_constraint.replace('_', ' ').capitalize()
    sheet_title += f"{model_type_text} Model Results - {objective_function_text} with {constraint_text}"
    return sheet_title
